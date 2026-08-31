from fastapi import APIRouter, Depends, HTTPException, Request, UploadFile, File, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import text
from typing import Dict, Any, Optional, List
from datetime import datetime
import uuid
import csv
import io

from database import get_db
from core.metadata_engine import MetadataEngine
from core.dynamic_verification import DynamicVerificationEngine
from core.auth import get_current_user, optional_get_current_user, require_permission
from core.audit import log_dynamic_audit
from core.errors import secure_db_error, ErrorCodes, create_error_response
from core.rate_limit import read_limiter, write_limiter
from core.query_parser import QueryParser, QueryParseError
from core.event_bus import EventBus


router = APIRouter(
    prefix="/api/v1/dynamic",
    tags=["Dynamic CRUD"]
)

BULK_MAX_RECORDS = 500


def get_verification_engine(
    entity_code: str,
    db: Session = Depends(get_db)
):
    return DynamicVerificationEngine(db, entity_code)


# ──────────────────────────────────────────────────────────────
# GET SCHEMA
# ──────────────────────────────────────────────────────────────

@router.get("/entities/{entity_code}/schema",
            dependencies=[
                Depends(require_permission("dynamic", "read")),
                Depends(read_limiter.check)
            ])
async def get_schema(
    entity_code: str,
    db: Session = Depends(get_db),
    verification: DynamicVerificationEngine = Depends(
        get_verification_engine
    )
):
    if not verification.entity_exists():
        raise HTTPException(
            status_code=404,
            detail=f"الكيان '{entity_code}' غير موجود"
        )

    table_error = verification.validate_table_mapping()
    if table_error:
        raise HTTPException(
            status_code=400,
            detail=table_error
        )

    engine = MetadataEngine(db)

    return {
        "status": "success",
        "data": engine.get_full_schema(entity_code),
        "table": verification.entity_meta.get("table_mapping"),
        "tenant_capability": verification.tenant_capability,
        "real_columns": verification.get_table_columns(),
    }


# ──────────────────────────────────────────────────────────────
# GET RECORDS (with P7 Query Engine)
# ──────────────────────────────────────────────────────────────

@router.get("/entities/{entity_code}/records",
            dependencies=[
                Depends(require_permission("dynamic", "read")),
                Depends(read_limiter.check)
            ])
async def list_records(
    entity_code: str,
    filters: Optional[str] = None,
    sort: Optional[str] = None,
    limit: int = 100,
    offset: int = 0,
    include: Optional[str] = None,
    include_deleted: bool = False,
    db: Session = Depends(get_db),
    current_user: Optional[dict] = Depends(optional_get_current_user)
):
    # Entity lookup via raw SQL (avoid DynamicVerificationEngine inspector)
    ent_row = db.execute(
        text(
            "SELECT id, table_mapping FROM dbp_entities "
            "WHERE code = :code"
        ),
        {"code": entity_code}
    ).fetchone()

    if not ent_row:
        raise HTTPException(
            status_code=404,
            detail=f"الكيان '{entity_code}' غير موجود"
        )

    entity_id = ent_row[0]
    table_mapping = ent_row[1]

    if not table_mapping:
        raise HTTPException(
            status_code=400,
            detail="Entity has no table mapping"
        )

    # Validate table name
    import re
    if not re.match(r'^[a-z0-9_]+$', table_mapping):
        raise HTTPException(
            status_code=400,
            detail="Invalid table name"
        )

    # Get real columns via raw SQL
    col_rows = db.execute(
        text(
            "SELECT column_name FROM information_schema.columns "
            "WHERE table_name = :tname ORDER BY ordinal_position"
        ),
        {"tname": table_mapping}
    ).fetchall()

    real_columns = {row[0].lower(): row[0] for row in col_rows}
    has_tenant = "tenant_id" in real_columns

    # Parse and validate query parameters
    try:
        parser = QueryParser(real_columns)
        query_filter = parser.parse_query(
            filters_str=filters,
            sort_str=sort,
            limit=limit,
            offset=offset
        )
    except QueryParseError as e:
        return create_error_response(
            status_code=400,
            code=ErrorCodes.VALIDATION_ERROR,
            details=[{"message": str(e), "message_en": str(e)}]
        )

    table_name = table_mapping
    where_clauses = []
    params = {}

    # P10: include_deleted access control
    has_deleted_at_col = "deleted_at" in {
        c.lower() for c in real_columns.keys()
    }
    if include_deleted and has_deleted_at_col:
        # Only admin can view deleted records
        if not current_user:
            raise HTTPException(
                status_code=401,
                detail="Authentication required"
            )
        # Check for admin: *:* permission or "admin" role
        roles = current_user.get("roles", [])
        is_admin = False
        for r in roles:
            if isinstance(r, dict) and r.get("permission") == "*:*":
                is_admin = True
                break
            if r == "admin":
                is_admin = True
                break
        # Also check permissions list
        perms = current_user.get("permissions", [])
        if "*:*" in perms:
            is_admin = True
        if not is_admin:
            raise HTTPException(
                status_code=403,
                detail="include_deleted requires admin role"
            )
        # Admin sees all records (including deleted)
    elif has_deleted_at_col:
        # Default: exclude soft-deleted records
        where_clauses.append("deleted_at IS NULL")

    # Tenant scope (NEVER from user filters)
    if has_tenant:
        # SCOPED: authenticated tenant required
        if not current_user:
            raise HTTPException(
                status_code=401,
                detail="Authentication required for SCOPED entity"
            )

        auth_tenant_id = current_user["tenant_id"]
        where_clauses.append("tenant_id = :tenant_id")
        params["tenant_id"] = auth_tenant_id
        tenant_applied = True
        effective_tenant = auth_tenant_id

    else:
        tenant_applied = False
        effective_tenant = None

    # Add user filters (parameterized)
    user_where, user_params = parser.build_where_clause(query_filter.filters)
    if user_where:
        where_clauses.append(user_where)
        params.update(user_params)

    # Build WHERE clause
    where_sql = " AND ".join(where_clauses) if where_clauses else "1=1"

    # Build ORDER BY clause
    order_sql = parser.build_order_clause(query_filter.sorts)
    if not order_sql:
        order_sql = "ORDER BY created_at DESC"  # Default sort

    # Get total count (with filters applied)
    count_query = text(
        f"SELECT COUNT(*) FROM {table_name} WHERE {where_sql}"
    )
    total = db.execute(count_query, params).scalar()

    # Get records with pagination
    limit = query_filter.limit
    offset = query_filter.offset

    query = text(
        f"SELECT * FROM {table_name} "
        f"WHERE {where_sql} "
        f"{order_sql} "
        f"LIMIT :limit OFFSET :offset"
    )
    params["limit"] = limit
    params["offset"] = offset

    result = db.execute(query, params)

    # Calculate pagination
    has_next = (offset + limit) < total

    records = [dict(row._mapping) for row in result]

    # Resolve ?include= relationships (raw SQL only, no ORM)
    if include:
        include_list = [
            x.strip() for x in include.split(",") if x.strip()
        ]

        # Get entity ID via raw SQL
        ent_row = db.execute(
            text("SELECT id FROM dbp_entities WHERE code = :code"),
            {"code": entity_code}
        ).fetchone()

        if ent_row:
            entity_id = ent_row[0]

            for inc in include_list:
                # Get relationship definition via raw SQL
                rel_row = db.execute(
                    text(
                        "SELECT target_entity_code, relationship_type, "
                        "source_column, target_column, lookup_field, "
                        "tenant_scope, junction_table, "
                        "junction_source_col, junction_target_col "
                        "FROM dbp_relationships "
                        "WHERE entity_id = :eid AND code = :code"
                    ),
                    {"eid": entity_id, "code": inc}
                ).fetchone()

                if not rel_row:
                    continue  # Unknown relationship, skip

                rel_data = {
                    "target_entity_code": rel_row[0],
                    "relationship_type": rel_row[1],
                    "source_column": rel_row[2],
                    "target_column": rel_row[3],
                    "lookup_field": rel_row[4],
                    "tenant_scope": rel_row[5],
                    "junction_table": rel_row[6] or "",
                    "junction_source_col": rel_row[7] or "",
                    "junction_target_col": rel_row[8] or "",
                }

                # Get target table via raw SQL
                tgt_row = db.execute(
                    text(
                        "SELECT table_mapping FROM dbp_entities "
                        "WHERE code = :code"
                    ),
                    {"code": rel_data["target_entity_code"]}
                ).fetchone()

                if not tgt_row or not tgt_row[0]:
                    continue

                target_table = tgt_row[0]

                for record in records:
                    source_value = record.get(rel_data["source_column"])
                    if source_value is None:
                        record[inc] = []
                        continue

                    if rel_data["relationship_type"] == "many_to_many":
                        junction = rel_data["junction_table"]
                        j_src = rel_data["junction_source_col"]
                        j_tgt = rel_data["junction_target_col"]
                        if not junction or not j_src or not j_tgt:
                            record[inc] = []
                            continue
                        q = text(
                            f"SELECT t.* FROM {target_table} t "
                            f"INNER JOIN {junction} j "
                            f"ON j.{j_tgt} = t.{rel_data['target_column']} "
                            f"WHERE j.{j_src} = :sv LIMIT 100"
                        )
                        rows = db.execute(q, {"sv": source_value}).fetchall()
                        if rows:
                            cols = [c for c in rows[0]._mapping.keys() if c != "tenant_id"]
                            record[inc] = [
                                {col: row._mapping[col] for col in cols}
                                for row in rows
                            ]
                        else:
                            record[inc] = []

                    elif rel_data["relationship_type"] == "lookup":
                        lookup_col = rel_data["lookup_field"]
                        where_parts = [
                            f"{rel_data['target_column']} = :sv"
                        ]
                        params: dict = {"sv": source_value}
                        if rel_data["tenant_scope"] and effective_tenant:
                            where_parts.append("tenant_id = :tid")
                            params["tid"] = effective_tenant
                        where_sql = " AND ".join(where_parts)
                        q = text(
                            f"SELECT id, {lookup_col} FROM {target_table} "
                            f"WHERE {where_sql} "
                            f"ORDER BY {lookup_col} ASC LIMIT 100"
                        )
                        rows = db.execute(q, params).fetchall()
                        record[inc] = [
                            {"id": str(r[0]), "label": str(r[1])}
                            for r in rows
                        ]

                    else:  # one_to_many, many_to_one
                        where_parts = [
                            f"{rel_data['target_column']} = :sv"
                        ]
                        params = {"sv": source_value}
                        if rel_data["tenant_scope"] and effective_tenant:
                            where_parts.append("tenant_id = :tid")
                            params["tid"] = effective_tenant
                        where_sql = " AND ".join(where_parts)
                        q = text(
                            f"SELECT * FROM {target_table} "
                            f"WHERE {where_sql} LIMIT 100"
                        )
                        rows = db.execute(q, params).fetchall()
                        if rows:
                            cols = [
                                c for c in rows[0]._mapping.keys()
                                if c != "tenant_id"
                            ]
                            record[inc] = [
                                {col: row._mapping[col] for col in cols}
                                for row in rows
                            ]
                        else:
                            record[inc] = []

    return {
        "status": "success",
        "data": records,
        "count": len(records),
        "tenant_applied": tenant_applied,
        "effective_tenant": effective_tenant,
        "pagination": {
            "total": total,
            "limit": limit,
            "offset": offset,
            "has_next": has_next,
        },
    }


# ──────────────────────────────────────────────────────────────
# CREATE SINGLE RECORD
# ──────────────────────────────────────────────────────────────

@router.post("/entities/{entity_code}/records",
             dependencies=[
                 Depends(require_permission("dynamic", "create")),
                 Depends(write_limiter.check)
             ])
async def create_record(
    entity_code: str,
    payload: Dict[str, Any],
    request: Request,
    db: Session = Depends(get_db),
    verification: DynamicVerificationEngine = Depends(
        get_verification_engine
    ),
    current_user: Optional[dict] = Depends(optional_get_current_user)
):
    if not verification.entity_exists():
        raise HTTPException(
            status_code=404,
            detail=f"الكيان '{entity_code}' غير موجود"
        )

    table_error = verification.validate_table_mapping()
    if table_error:
        raise HTTPException(
            status_code=400,
            detail=table_error
        )

    try:
        verification.validate_required_fields(payload)
        verification.validate_enum_fields(payload)
    except ValueError as e:
        # Validation error - return structured response
        return create_error_response(
            status_code=400,
            code=ErrorCodes.VALIDATION_ERROR,
            details=[{"message": str(e), "message_en": str(e)}]
        )

    not_null_errors = verification.validate_not_null_columns(
        payload,
        exclude_cols=["tenant_id"]
    )

    if not_null_errors:
        return create_error_response(
            status_code=400,
            code=ErrorCodes.NOT_NULL_VIOLATION,
            details=[{"message": e, "message_en": e} for e in not_null_errors]
        )

    # Determine effective tenant for duplicate check
    effective_tenant = None
    if verification.has_tenant_id_column():
        if current_user:
            effective_tenant = current_user["tenant_id"]

    duplicate_errors = verification.check_duplicate_by_unique_fields(
        payload,
        tenant_id=effective_tenant
    )

    if duplicate_errors:
        return create_error_response(
            status_code=409,
            code=ErrorCodes.DUPLICATE_RECORD,
            details=[{"message": e, "message_en": e} for e in duplicate_errors]
        )

    table_name = verification.entity_meta["table_mapping"]

    # Generate PK based on table schema
    clean_data = {verification.get_pk_column(): verification.generate_pk_value()}

    # Determine effective tenant
    effective_tenant = None

    if verification.has_tenant_id_column():
        # SCOPED: authenticated tenant required
        if not current_user:
            raise HTTPException(
                status_code=401,
                detail="Authentication required for SCOPED entity"
            )

        effective_tenant = current_user["tenant_id"]

        # Force authenticated tenant — ignore payload.tenant_id
        clean_data["tenant_id"] = effective_tenant

    pk_col = verification.get_pk_column()
    for key, value in payload.items():
        if key == pk_col:
            continue  # Skip PK column
        if key == "tenant_id":
            # tenant_id already handled above
            continue
        if verification.check_column_exists(key):
            clean_data[key] = value

    cols = ", ".join(clean_data.keys())
    placeholders = ", ".join(
        [f":{k}" for k in clean_data.keys()]
    )

    query = text(
        f"""
        INSERT INTO {table_name}
        ({cols})
        VALUES ({placeholders})
        RETURNING {pk_col}
        """
    )

    try:
        result = db.execute(query, clean_data)

        new_id = result.scalar()

        # Audit: success (before commit, part of same transaction)
        log_dynamic_audit(
            db=db,
            tenant_id=effective_tenant or "unknown",
            user_id=current_user.get("id") if current_user else None,
            user_email=current_user.get("email") if current_user else None,
            action="create",
            entity_code=entity_code,
            record_id=new_id,
            new_values=clean_data,
            ip_address=request.client.host if request.client else None,
            user_agent=request.headers.get("user-agent") if request else None,
            status="success",
        )

        EventBus(db).emit("record.created", entity_code, tenant_id=effective_tenant, user_id=current_user.get("id") if current_user else None, record_id=new_id, payload=clean_data)

        db.commit()

        return {
            "status": "success",
            "id": new_id,
            "message": "تم إنشاؤه بنجاح",
            "tenant_capability": verification.tenant_capability,
            "effective_tenant": effective_tenant,
        }

    except Exception as e:
        db.rollback()

        # Audit: failure
        log_dynamic_audit(
            db=db,
            tenant_id=effective_tenant or "unknown",
            user_id=current_user.get("id") if current_user else None,
            user_email=current_user.get("email") if current_user else None,
            action="create",
            entity_code=entity_code,
            new_values=clean_data,
            ip_address=request.client.host if request.client else None,
            user_agent=request.headers.get("user-agent") if request else None,
            status="failure",
            error_message=str(e),
        )

        # Secure error: never expose raw exception
        return secure_db_error(e)


# ──────────────────────────────────────────────────────────────
# P8.1 — BULK CREATE (MUST be before /records/{record_id})
# ──────────────────────────────────────────────────────────────

@router.post("/entities/{entity_code}/records/bulk",
             dependencies=[
                 Depends(require_permission("dynamic", "create")),
                 Depends(write_limiter.check)
             ])
async def bulk_create_records(
    entity_code: str,
    payload: Dict[str, Any],
    request: Request,
    db: Session = Depends(get_db),
    verification: DynamicVerificationEngine = Depends(
        get_verification_engine
    ),
    current_user: Optional[dict] = Depends(optional_get_current_user)
):
    """
    Create multiple records in a single request.

    Request body:
        { "records": [ {field: value}, ... ] }

    Rules:
        - Max 500 records per request
        - Atomic transaction: all succeed or all fail
        - Same validation pipeline as single create
        - Per-record audit logging
        - Tenant isolation enforced
    """
    if not verification.entity_exists():
        raise HTTPException(
            status_code=404,
            detail=f"الكيان '{entity_code}' غير موجود"
        )

    table_error = verification.validate_table_mapping()
    if table_error:
        raise HTTPException(
            status_code=400,
            detail=table_error
        )

    # Extract records from payload
    records = payload.get("records")
    if not records or not isinstance(records, list):
        return create_error_response(
            status_code=400,
            code=ErrorCodes.VALIDATION_ERROR,
            details=[{"message": "حقل 'records' مطلوب ويجب أن يكون مصفوفة",
                       "message_en": "Field 'records' is required and must be an array"}]
        )

    if len(records) == 0:
        return create_error_response(
            status_code=400,
            code=ErrorCodes.VALIDATION_ERROR,
            details=[{"message": "القائمة فارغة",
                       "message_en": "Records list is empty"}]
        )

    if len(records) > BULK_MAX_RECORDS:
        return create_error_response(
            status_code=400,
            code=ErrorCodes.VALIDATION_ERROR,
            details=[{"message": f"الحد الأقصى {BULK_MAX_RECORDS} سجل في الطلب",
                       "message_en": f"Maximum {BULK_MAX_RECORDS} records per request"}]
        )

    table_name = verification.entity_meta["table_mapping"]
    pk_col = verification.get_pk_column()

    # Determine effective tenant (same logic as single create)
    effective_tenant = None
    if verification.has_tenant_id_column():
        if not current_user:
            raise HTTPException(
                status_code=401,
                detail="Authentication required for SCOPED entity"
            )
        effective_tenant = current_user["tenant_id"]

    # Pre-validate all records before inserting
    collected_errors: List[Dict[str, Any]] = []
    clean_records: List[Dict[str, Any]] = []

    for idx, record in enumerate(records):
        # Must be a dict
        if not isinstance(record, dict):
            collected_errors.append({
                "index": idx,
                "code": ErrorCodes.VALIDATION_ERROR,
                "message": f"السجل #{idx} يجب أن يكون كائن (object)",
                "message_en": f"Record #{idx} must be an object",
            })
            continue

        # Schema validation
        try:
            verification.validate_required_fields(record)
            verification.validate_enum_fields(record)
        except ValueError as e:
            collected_errors.append({
                "index": idx,
                "code": ErrorCodes.VALIDATION_ERROR,
                "message": str(e),
                "message_en": str(e),
            })
            continue

        # NOT NULL validation
        not_null_errors = verification.validate_not_null_columns(
            record,
            exclude_cols=["tenant_id"]
        )
        if not_null_errors:
            collected_errors.append({
                "index": idx,
                "code": ErrorCodes.NOT_NULL_VIOLATION,
                "message": "; ".join(not_null_errors),
                "message_en": "; ".join(not_null_errors),
            })
            continue

        # Duplicate check
        dup_errors = verification.check_duplicate_by_unique_fields(
            record,
            tenant_id=effective_tenant
        )
        if dup_errors:
            collected_errors.append({
                "index": idx,
                "code": ErrorCodes.DUPLICATE_RECORD,
                "message": "; ".join(dup_errors),
                "message_en": "; ".join(dup_errors),
            })
            continue

        # Build clean data
        clean_data = {pk_col: verification.generate_pk_value()}

        if verification.has_tenant_id_column():
            clean_data["tenant_id"] = effective_tenant

        for key, value in record.items():
            if key == pk_col:
                continue
            if key == "tenant_id":
                continue
            if verification.check_column_exists(key):
                clean_data[key] = value

        clean_records.append(clean_data)

    # If there are validation errors, return them (atomic: don't insert any)
    if collected_errors:
        return create_error_response(
            status_code=400,
            code=ErrorCodes.VALIDATION_ERROR,
            details=collected_errors
        )

    # All records validated — insert in a single transaction
    created_ids = []
    try:
        for clean_data in clean_records:
            cols = ", ".join(clean_data.keys())
            placeholders = ", ".join(
                [f":{k}" for k in clean_data.keys()]
            )

            query = text(
                f"""
                INSERT INTO {table_name}
                ({cols})
                VALUES ({placeholders})
                RETURNING {pk_col}
                """
            )

            result = db.execute(query, clean_data)
            new_id = result.scalar()
            created_ids.append(new_id)

            # Per-record audit
            log_dynamic_audit(
                db=db,
                tenant_id=effective_tenant or "unknown",
                user_id=current_user.get("id") if current_user else None,
                user_email=current_user.get("email") if current_user else None,
                action="bulk_create",
                entity_code=entity_code,
                record_id=new_id,
                new_values=clean_data,
                ip_address=request.client.host if request.client else None,
                user_agent=request.headers.get("user-agent") if request else None,
                status="success",
            )

        EventBus(db).emit("record.bulk_created", entity_code, tenant_id=effective_tenant, user_id=current_user.get("id") if current_user else None, payload={"created_ids": created_ids, "count": len(created_ids)})

        db.commit()

        return {
            "status": "success",
            "created": created_ids,
            "count": len(created_ids),
            "tenant_capability": verification.tenant_capability,
            "effective_tenant": effective_tenant,
        }

    except Exception as e:
        db.rollback()

        # Audit: bulk failure
        log_dynamic_audit(
            db=db,
            tenant_id=effective_tenant or "unknown",
            user_id=current_user.get("id") if current_user else None,
            user_email=current_user.get("email") if current_user else None,
            action="bulk_create",
            entity_code=entity_code,
            new_values={"attempted_count": len(clean_records)},
            ip_address=request.client.host if request.client else None,
            user_agent=request.headers.get("user-agent") if request else None,
            status="failure",
            error_message=str(e),
        )

        return secure_db_error(e)


# ──────────────────────────────────────────────────────────────
# P8.2 — BULK UPDATE (MUST be before /records/{record_id})
# ──────────────────────────────────────────────────────────────

@router.put("/entities/{entity_code}/records/bulk",
            dependencies=[
                Depends(require_permission("dynamic", "update")),
                Depends(write_limiter.check)
            ])
async def bulk_update_records(
    entity_code: str,
    payload: Dict[str, Any],
    request: Request,
    db: Session = Depends(get_db),
    verification: DynamicVerificationEngine = Depends(
        get_verification_engine
    ),
    current_user: Optional[dict] = Depends(optional_get_current_user)
):
    """
    Update multiple records in a single request.

    Request body:
        { "records": [{"id": "...", "fields": {field: value}}, ...] }

    Rules:
        - Max 500 records per request
        - Atomic transaction: all succeed or all fail
        - Tenant isolation enforced per record
        - Per-record audit logging
    """
    if not verification.entity_exists():
        raise HTTPException(
            status_code=404,
            detail=f"الكيان '{entity_code}' غير موجود"
        )

    table_error = verification.validate_table_mapping()
    if table_error:
        raise HTTPException(
            status_code=400,
            detail=table_error
        )

    records = payload.get("records")
    if not records or not isinstance(records, list):
        return create_error_response(
            status_code=400,
            code=ErrorCodes.VALIDATION_ERROR,
            details=[{"message": "حقل 'records' مطلوب ويجب أن يكون مصفوفة",
                       "message_en": "Field 'records' is required and must be an array"}]
        )

    if len(records) == 0:
        return create_error_response(
            status_code=400,
            code=ErrorCodes.VALIDATION_ERROR,
            details=[{"message": "القائمة فارغة",
                       "message_en": "Records list is empty"}]
        )

    if len(records) > BULK_MAX_RECORDS:
        return create_error_response(
            status_code=400,
            code=ErrorCodes.VALIDATION_ERROR,
            details=[{"message": f"الحد الأقصى {BULK_MAX_RECORDS} سجل في الطلب",
                       "message_en": f"Maximum {BULK_MAX_RECORDS} records per request"}]
        )

    table_name = verification.entity_meta["table_mapping"]
    pk_col = verification.get_pk_column()

    # Determine effective tenant
    effective_tenant = None
    if verification.has_tenant_id_column():
        if not current_user:
            raise HTTPException(
                status_code=401,
                detail="Authentication required for SCOPED entity"
            )
        effective_tenant = current_user["tenant_id"]

    # Pre-validate all records
    collected_errors: List[Dict[str, Any]] = []
    update_specs: List[Dict[str, Any]] = []

    for idx, record in enumerate(records):
        if not isinstance(record, dict):
            collected_errors.append({
                "index": idx,
                "code": ErrorCodes.VALIDATION_ERROR,
                "message": f"السجل #{idx} يجب أن يكون كائن (object)",
                "message_en": f"Record #{idx} must be an object",
            })
            continue

        record_id = record.get("id")
        if not record_id:
            collected_errors.append({
                "index": idx,
                "code": ErrorCodes.VALIDATION_ERROR,
                "message": f"السجل #{idx}: حقل 'id' مطلوب",
                "message_en": f"Record #{idx}: 'id' field is required",
            })
            continue

        fields = record.get("fields")
        if not fields or not isinstance(fields, dict):
            collected_errors.append({
                "index": idx,
                "code": ErrorCodes.VALIDATION_ERROR,
                "message": f"السجل #{idx}: حقل 'fields' مطلوب",
                "message_en": f"Record #{idx}: 'fields' field is required",
            })
            continue

        # Filter valid columns (skip id, tenant_id)
        valid_cols = [
            k for k in fields.keys()
            if k != pk_col
            and k != "tenant_id"
            and verification.check_column_exists(k)
        ]

        if not valid_cols:
            collected_errors.append({
                "index": idx,
                "code": ErrorCodes.VALIDATION_ERROR,
                "message": f"السجل #{idx}: لا توجد أعمدة صالحة للتحديث",
                "message_en": f"Record #{idx}: no valid columns to update",
            })
            continue

        update_specs.append({
            "id": record_id,
            "fields": fields,
            "valid_cols": valid_cols,
            "index": idx,
        })

    if collected_errors:
        return create_error_response(
            status_code=400,
            code=ErrorCodes.VALIDATION_ERROR,
            details=collected_errors
        )

    # Execute updates atomically
    updated_count = 0
    try:
        for spec in update_specs:
            update_params = {"id": spec["id"]}

            set_parts = []
            for key in spec["valid_cols"]:
                set_parts.append(f"{key} = :{key}")
                update_params[key] = spec["fields"][key]

            set_clause = ", ".join(set_parts)
            where_clause = f"WHERE {pk_col} = :id"

            # P10: Exclude soft-deleted records from update
            has_deleted_bulk = verification.real_columns.get("deleted_at") is not None
            if has_deleted_bulk:
                where_clause += " AND deleted_at IS NULL"

            if verification.has_tenant_id_column():
                where_clause += " AND tenant_id = :tenant_filter"
                update_params["tenant_filter"] = effective_tenant

            query = text(
                f"""
                UPDATE {table_name}
                SET {set_clause}
                {where_clause}
                """
            )

            result = db.execute(query, update_params)

            if result.rowcount == 0:
                collected_errors.append({
                    "index": spec["index"],
                    "code": ErrorCodes.RECORD_NOT_FOUND,
                    "message": f"السجل '{spec['id']}' غير موجود",
                    "message_en": f"Record '{spec['id']}' not found",
                })
                continue

            updated_count += 1

            # Per-record audit
            new_values = {k: spec["fields"][k] for k in spec["valid_cols"]}
            log_dynamic_audit(
                db=db,
                tenant_id=effective_tenant or "unknown",
                user_id=current_user.get("id") if current_user else None,
                user_email=current_user.get("email") if current_user else None,
                action="bulk_update",
                entity_code=entity_code,
                record_id=spec["id"],
                new_values=new_values,
                ip_address=request.client.host if request.client else None,
                user_agent=request.headers.get("user-agent") if request else None,
                status="success",
            )

        if collected_errors:
            db.rollback()
            return create_error_response(
                status_code=400,
                code=ErrorCodes.VALIDATION_ERROR,
                details=collected_errors
            )

        EventBus(db).emit("record.bulk_updated", entity_code, tenant_id=effective_tenant, user_id=current_user.get("id") if current_user else None, payload={"updated_count": updated_count})

        db.commit()

        return {
            "status": "success",
            "updated": updated_count,
            "tenant_capability": verification.tenant_capability,
            "effective_tenant": effective_tenant,
        }

    except Exception as e:
        db.rollback()

        log_dynamic_audit(
            db=db,
            tenant_id=effective_tenant or "unknown",
            user_id=current_user.get("id") if current_user else None,
            user_email=current_user.get("email") if current_user else None,
            action="bulk_update",
            entity_code=entity_code,
            new_values={"attempted_count": len(update_specs)},
            ip_address=request.client.host if request.client else None,
            user_agent=request.headers.get("user-agent") if request else None,
            status="failure",
            error_message=str(e),
        )

        return secure_db_error(e)


# ──────────────────────────────────────────────────────────────
# P8.3 — BULK DELETE (MUST be before /records/{record_id})
# ──────────────────────────────────────────────────────────────

@router.delete("/entities/{entity_code}/records/bulk",
               dependencies=[
                   Depends(require_permission("dynamic", "delete")),
                   Depends(write_limiter.check)
               ])
async def bulk_delete_records(
    entity_code: str,
    payload: Dict[str, Any],
    request: Request,
    db: Session = Depends(get_db),
    current_user: Optional[dict] = Depends(optional_get_current_user)
):
    """
    Delete multiple records in a single request.
    """
    # Entity lookup via raw SQL (avoid DynamicVerificationEngine)
    ent_row = db.execute(
        text("SELECT id, table_mapping FROM dbp_entities WHERE code = :code"),
        {"code": entity_code}
    ).fetchone()
    if not ent_row:
        raise HTTPException(status_code=404, detail=f"الكيان '{entity_code}' غير موجود")
    entity_id = ent_row[0]
    table_mapping = ent_row[1]
    if not table_mapping:
        raise HTTPException(status_code=400, detail="Entity has no table mapping")

    import re
    if not re.match(r'^[a-z0-9_]+$', table_mapping):
        raise HTTPException(status_code=400, detail="Invalid table name")

    table_name = table_mapping

    # Get real columns
    col_rows = db.execute(
        text("SELECT column_name FROM information_schema.columns WHERE table_name = :tname"),
        {"tname": table_name}
    ).fetchall()
    real_columns = {row[0].lower(): row[0] for row in col_rows}
    has_tenant = "tenant_id" in real_columns
    pk_col = "id"

    ids = payload.get("ids")
    if not ids or not isinstance(ids, list):
        return create_error_response(
            status_code=400,
            code=ErrorCodes.VALIDATION_ERROR,
            details=[{"message": "حقل 'ids' مطلوب ويجب أن يكون مصفوفة",
                       "message_en": "Field 'ids' is required and must be an array"}]
        )

    if len(ids) == 0:
        return create_error_response(
            status_code=400,
            code=ErrorCodes.VALIDATION_ERROR,
            details=[{"message": "القائمة فارغة",
                       "message_en": "IDs list is empty"}]
        )

    if len(ids) > BULK_MAX_RECORDS:
        return create_error_response(
            status_code=400,
            code=ErrorCodes.VALIDATION_ERROR,
            details=[{"message": f"الحد الأقصى {BULK_MAX_RECORDS} سجل في الطلب",
                       "message_en": f"Maximum {BULK_MAX_RECORDS} records per request"}]
        )

    # Determine effective tenant
    effective_tenant = None
    if has_tenant:
        if not current_user:
            raise HTTPException(
                status_code=401,
                detail="Authentication required for SCOPED entity"
            )
        effective_tenant = current_user["tenant_id"]

    # P10: Check if table has deleted_at column
    has_deleted = "deleted_at" in real_columns
    user_id = current_user.get("id") if current_user else None
    now = datetime.utcnow()

    # Execute deletes atomically
    deleted_count = 0
    collected_errors: List[Dict[str, Any]] = []

    try:
        for idx, record_id in enumerate(ids):
            # Get old values for audit
            old_values = None
            try:
                if has_tenant:
                    old_q = text(f"SELECT * FROM {table_name} WHERE {pk_col} = :id AND tenant_id = :tid")
                    old_r = db.execute(old_q, {"id": record_id, "tid": effective_tenant})
                else:
                    old_q = text(f"SELECT * FROM {table_name} WHERE {pk_col} = :id")
                    old_r = db.execute(old_q, {"id": record_id})
                old_row = old_r.fetchone()
                if old_row:
                    old_values = dict(old_row._mapping)
                    # Skip already soft-deleted records
                    if has_deleted and old_values.get("deleted_at") is not None:
                        collected_errors.append({
                            "index": idx,
                            "code": ErrorCodes.RECORD_NOT_FOUND,
                            "message": f"السجل '{record_id}' غير موجود",
                            "message_en": f"Record '{record_id}' not found",
                        })
                        continue
            except Exception:
                pass

            if has_deleted:
                # Soft delete
                if has_tenant:
                    query = text(
                        f"UPDATE {table_name} SET deleted_at = :deleted_at, deleted_by = :deleted_by "
                        f"WHERE {pk_col} = :id AND tenant_id = :tenant_id AND deleted_at IS NULL"
                    )
                    result = db.execute(query, {
                        "deleted_at": now, "deleted_by": user_id,
                        "id": record_id, "tenant_id": effective_tenant
                    })
                else:
                    query = text(
                        f"UPDATE {table_name} SET deleted_at = :deleted_at, deleted_by = :deleted_by "
                        f"WHERE {pk_col} = :id AND deleted_at IS NULL"
                    )
                    result = db.execute(query, {
                        "deleted_at": now, "deleted_by": user_id, "id": record_id
                    })
            else:
                # Physical delete fallback
                if has_tenant:
                    query = text(f"DELETE FROM {table_name} WHERE {pk_col} = :id AND tenant_id = :tenant_id")
                    result = db.execute(query, {"id": record_id, "tenant_id": effective_tenant})
                else:
                    query = text(f"DELETE FROM {table_name} WHERE {pk_col} = :id")
                    result = db.execute(query, {"id": record_id})

            if result.rowcount == 0:
                collected_errors.append({
                    "index": idx,
                    "code": ErrorCodes.RECORD_NOT_FOUND,
                    "message": f"السجل '{record_id}' غير موجود",
                    "message_en": f"Record '{record_id}' not found",
                })
                continue

            deleted_count += 1

            # Per-record audit
            log_dynamic_audit(
                db=db,
                tenant_id=effective_tenant or "unknown",
                user_id=user_id,
                user_email=current_user.get("email") if current_user else None,
                action="bulk_delete",
                entity_code=entity_code,
                record_id=record_id,
                old_values=old_values,
                ip_address=request.client.host if request.client else None,
                user_agent=request.headers.get("user-agent") if request else None,
                status="success",
            )

        if collected_errors:
            db.rollback()
            return create_error_response(
                status_code=400,
                code=ErrorCodes.VALIDATION_ERROR,
                details=collected_errors
            )

        EventBus(db).emit("record.bulk_deleted", entity_code, tenant_id=effective_tenant, user_id=current_user.get("id") if current_user else None, payload={"deleted_count": deleted_count, "ids": ids})

        db.commit()

        return {
            "status": "success",
            "deleted": deleted_count,
            "effective_tenant": effective_tenant,
        }

    except Exception as e:
        db.rollback()

        log_dynamic_audit(
            db=db,
            tenant_id=effective_tenant or "unknown",
            user_id=current_user.get("id") if current_user else None,
            user_email=current_user.get("email") if current_user else None,
            action="bulk_delete",
            entity_code=entity_code,
            new_values={"attempted_ids": ids},
            ip_address=request.client.host if request.client else None,
            user_agent=request.headers.get("user-agent") if request else None,
            status="failure",
            error_message=str(e),
        )

        return secure_db_error(e)


# ──────────────────────────────────────────────────────────────
# UPDATE SINGLE RECORD
# ──────────────────────────────────────────────────────────────

@router.put("/entities/{entity_code}/records/{record_id}",
            dependencies=[
                Depends(require_permission("dynamic", "update")),
                Depends(write_limiter.check)
            ])
async def update_record(
    entity_code: str,
    record_id: str,
    payload: Dict[str, Any],
    request: Request,
    db: Session = Depends(get_db),
    verification: DynamicVerificationEngine = Depends(
        get_verification_engine
    ),
    current_user: Optional[dict] = Depends(optional_get_current_user)
):
    if not verification.entity_exists():
        raise HTTPException(
            status_code=404,
            detail=f"الكيان '{entity_code}' غير موجود"
        )

    table_error = verification.validate_table_mapping()
    if table_error:
        raise HTTPException(
            status_code=400,
            detail=table_error
        )

    valid_cols = [
        k for k in payload.keys()
        if k != "id"
        and k != "tenant_id"
        and verification.check_column_exists(k)
    ]

    if not valid_cols:
        raise HTTPException(
            status_code=400,
            detail="لا توجد أعمدة صالحة للتحديث"
        )

    table_name = verification.entity_meta["table_mapping"]
    update_params = {"id": record_id}

    set_parts = []
    for key in valid_cols:
        set_parts.append(f"{key} = :{key}")
        update_params[key] = payload[key]

    set_clause = ", ".join(set_parts)
    where_clause = "WHERE id = :id"

    # P10: Exclude soft-deleted records from update
    has_deleted = verification.real_columns.get("deleted_at") is not None
    if has_deleted:
        where_clause += " AND deleted_at IS NULL"

    if verification.has_tenant_id_column():
        # SCOPED: authenticated tenant required
        if not current_user:
            raise HTTPException(
                status_code=401,
                detail="Authentication required for SCOPED entity"
            )

        auth_tenant_id = current_user["tenant_id"]

        # Add tenant to WHERE clause for isolation
        where_clause += " AND tenant_id = :tenant_filter"
        update_params["tenant_filter"] = auth_tenant_id

    # Build new_values for audit (only the updated columns)
    new_values = {k: v for k, v in update_params.items() if k not in ("id", "tenant_filter")}

    # Get old values for audit (before update)
    auth_tenant_id_audit = current_user.get("tenant_id") if current_user else None
    old_values = None

    query = text(
        f"""
        UPDATE {table_name}
        SET {set_clause}
        {where_clause}
        """
    )

    try:
        result = db.execute(query, update_params)

        if result.rowcount == 0:
            raise HTTPException(
                status_code=404,
                detail="السجل غير موجود"
            )

        # Audit: success (before commit, part of same transaction)
        log_dynamic_audit(
            db=db,
            tenant_id=auth_tenant_id_audit or "unknown",
            user_id=current_user.get("id") if current_user else None,
            user_email=current_user.get("email") if current_user else None,
            action="update",
            entity_code=entity_code,
            record_id=record_id,
            old_values=old_values,
            new_values=new_values,
            ip_address=request.client.host if request.client else None,
            user_agent=request.headers.get("user-agent") if request else None,
            status="success",
        )

        EventBus(db).emit("record.updated", entity_code, tenant_id=auth_tenant_id_audit, user_id=current_user.get("id") if current_user else None, record_id=record_id, payload={"old": old_values, "new": new_values})

        db.commit()

        return {
            "status": "success",
            "message": "تم تحديث السجل بنجاح",
        }

    except HTTPException:
        raise

    except Exception as e:
        db.rollback()

        # Audit: failure
        log_dynamic_audit(
            db=db,
            tenant_id=auth_tenant_id_audit or "unknown",
            user_id=current_user.get("id") if current_user else None,
            user_email=current_user.get("email") if current_user else None,
            action="update",
            entity_code=entity_code,
            record_id=record_id,
            old_values=old_values,
            new_values=update_params,
            ip_address=request.client.host if request.client else None,
            user_agent=request.headers.get("user-agent") if request else None,
            status="failure",
            error_message=str(e),
        )

        # Secure error: never expose raw exception
        return secure_db_error(e)


# ──────────────────────────────────────────────────────────────
# DELETE SINGLE RECORD
# ──────────────────────────────────────────────────────────────

@router.delete("/entities/{entity_code}/records/{record_id}",
               dependencies=[
                   Depends(require_permission("dynamic", "delete")),
                   Depends(write_limiter.check)
               ])
async def delete_record(
    entity_code: str,
    record_id: str,
    request: Request,
    db: Session = Depends(get_db),
    current_user: Optional[dict] = Depends(optional_get_current_user)
):
    import re
    ent_row = db.execute(
        text("SELECT id, table_mapping FROM dbp_entities WHERE code = :code"),
        {"code": entity_code}
    ).fetchone()
    if not ent_row:
        raise HTTPException(status_code=404, detail=f"الكيان '{entity_code}' غير موجود")
    table_mapping = ent_row[1]
    if not table_mapping:
        raise HTTPException(status_code=400, detail="Entity has no table mapping")
    if not re.match(r'^[a-z0-9_]+$', table_mapping):
        raise HTTPException(status_code=400, detail="Invalid table name")

    table_name = table_mapping

    col_rows = db.execute(
        text("SELECT column_name FROM information_schema.columns WHERE table_name = :tname"),
        {"tname": table_name}
    ).fetchall()
    real_columns = {row[0].lower(): row[0] for row in col_rows}
    has_tenant = "tenant_id" in real_columns
    has_deleted = "deleted_at" in real_columns

    # Get old values for audit (before delete)
    auth_tenant_id_del = current_user.get("tenant_id") if current_user else None
    user_id = current_user.get("id") if current_user else None
    old_values = None
    try:
        if has_tenant:
            old_query = text(f"SELECT * FROM {table_name} WHERE id = :id AND tenant_id = :tenant_id")
            old_result = db.execute(old_query, {"id": record_id, "tenant_id": auth_tenant_id_del})
        else:
            old_query = text(f"SELECT * FROM {table_name} WHERE id = :id")
            old_result = db.execute(old_query, {"id": record_id})
        old_row = old_result.fetchone()
        if old_row:
            old_values = dict(old_row._mapping)
    except Exception:
        pass

    # P10: Soft delete if deleted_at column exists
    if has_deleted:
        # Check if already soft-deleted
        if has_tenant:
            check_q = text(
                f"SELECT deleted_at FROM {table_name} "
                f"WHERE id = :id AND tenant_id = :tenant_id"
            )
            check_r = db.execute(check_q, {"id": record_id, "tenant_id": auth_tenant_id_del})
        else:
            check_q = text(f"SELECT deleted_at FROM {table_name} WHERE id = :id")
            check_r = db.execute(check_q, {"id": record_id})
        check_row = check_r.fetchone()
        if check_row and check_row[0] is not None:
            raise HTTPException(status_code=404, detail="السجل غير موجود")

        # Soft delete: SET deleted_at, deleted_by
        now = datetime.utcnow()
        if has_tenant:
            if not current_user:
                raise HTTPException(status_code=401, detail="Authentication required for SCOPED entity")
            query = text(
                f"UPDATE {table_name} SET deleted_at = :deleted_at, deleted_by = :deleted_by "
                f"WHERE id = :id AND tenant_id = :tenant_id AND deleted_at IS NULL"
            )
            result = db.execute(query, {
                "deleted_at": now, "deleted_by": user_id,
                "id": record_id, "tenant_id": current_user["tenant_id"]
            })
        else:
            query = text(
                f"UPDATE {table_name} SET deleted_at = :deleted_at, deleted_by = :deleted_by "
                f"WHERE id = :id AND deleted_at IS NULL"
            )
            result = db.execute(query, {
                "deleted_at": now, "deleted_by": user_id, "id": record_id
            })
    else:
        # Physical delete fallback (no deleted_at column)
        if has_tenant:
            if not current_user:
                raise HTTPException(status_code=401, detail="Authentication required for SCOPED entity")
            query = text(f"DELETE FROM {table_name} WHERE id = :id AND tenant_id = :tenant_id")
            result = db.execute(query, {"id": record_id, "tenant_id": current_user["tenant_id"]})
        else:
            query = text(f"DELETE FROM {table_name} WHERE id = :id")
            result = db.execute(query, {"id": record_id})

    if result.rowcount == 0:
        log_dynamic_audit(
            db=db, tenant_id=auth_tenant_id_del or "unknown",
            user_id=user_id,
            user_email=current_user.get("email") if current_user else None,
            action="delete", entity_code=entity_code, record_id=record_id,
            old_values=old_values,
            ip_address=request.client.host if request.client else None,
            user_agent=request.headers.get("user-agent") if request else None,
            status="failure", error_message="Record not found",
        )
        db.commit()
        raise HTTPException(status_code=404, detail="السجل غير موجود")

    log_dynamic_audit(
        db=db, tenant_id=auth_tenant_id_del or "unknown",
        user_id=user_id,
        user_email=current_user.get("email") if current_user else None,
        action="delete", entity_code=entity_code, record_id=record_id,
        old_values=old_values,
        ip_address=request.client.host if request.client else None,
        user_agent=request.headers.get("user-agent") if request else None,
        status="success",
    )

    EventBus(db).emit("record.deleted", entity_code, tenant_id=auth_tenant_id_del, user_id=user_id, record_id=record_id, payload={"old": old_values})

    db.commit()

    return {
        "status": "success",
        "message": "تم حذف السجل بنجاح",
    }


# ──────────────────────────────────────────────────────────────
# P10 — RESTORE SOFT-DELETED RECORD
# ──────────────────────────────────────────────────────────────

@router.post("/entities/{entity_code}/records/{record_id}/restore",
             dependencies=[
                 Depends(require_permission("dynamic", "update")),
                 Depends(write_limiter.check)
             ])
async def restore_record(
    entity_code: str,
    record_id: str,
    request: Request,
    db: Session = Depends(get_db),
    verification: DynamicVerificationEngine = Depends(
        get_verification_engine
    ),
    current_user: Optional[dict] = Depends(optional_get_current_user)
):
    if not verification.entity_exists():
        raise HTTPException(status_code=404, detail=f"الكيان '{entity_code}' غير موجود")

    table_error = verification.validate_table_mapping()
    if table_error:
        raise HTTPException(status_code=400, detail=table_error)

    table_name = verification.entity_meta["table_mapping"]
    has_deleted = verification.real_columns.get("deleted_at") is not None

    if not has_deleted:
        raise HTTPException(status_code=400, detail="Entity does not support soft delete")

    if not current_user:
        raise HTTPException(status_code=401, detail="Authentication required")

    user_id = current_user.get("id")
    auth_tenant_id = current_user.get("tenant_id")

    # Check record exists and is soft-deleted
    if verification.has_tenant_id_column():
        check_q = text(
            f"SELECT deleted_at FROM {table_name} "
            f"WHERE id = :id AND tenant_id = :tenant_id"
        )
        check_r = db.execute(check_q, {"id": record_id, "tenant_id": auth_tenant_id})
    else:
        check_q = text(f"SELECT deleted_at FROM {table_name} WHERE id = :id")
        check_r = db.execute(check_q, {"id": record_id})

    check_row = check_r.fetchone()
    if not check_row:
        raise HTTPException(status_code=404, detail="السجل غير موجود")

    if check_row[0] is None:
        raise HTTPException(status_code=400, detail="السجل غير محذوف")

    # Restore: SET deleted_at = NULL, deleted_by = NULL
    if verification.has_tenant_id_column():
        restore_q = text(
            f"UPDATE {table_name} SET deleted_at = NULL, deleted_by = NULL "
            f"WHERE id = :id AND tenant_id = :tenant_id"
        )
        result = db.execute(restore_q, {"id": record_id, "tenant_id": auth_tenant_id})
    else:
        restore_q = text(
            f"UPDATE {table_name} SET deleted_at = NULL, deleted_by = NULL "
            f"WHERE id = :id"
        )
        result = db.execute(restore_q, {"id": record_id})

    if result.rowcount == 0:
        raise HTTPException(status_code=404, detail="السجل غير موجود")

    log_dynamic_audit(
        db=db, tenant_id=auth_tenant_id or "unknown",
        user_id=user_id,
        user_email=current_user.get("email") if current_user else None,
        action="restore", entity_code=entity_code, record_id=record_id,
        ip_address=request.client.host if request.client else None,
        user_agent=request.headers.get("user-agent") if request else None,
        status="success",
    )

    EventBus(db).emit("record.restored", entity_code, tenant_id=auth_tenant_id, user_id=user_id, record_id=record_id)

    db.commit()

    return {
        "status": "success",
        "message": "تم استرجاع السجل بنجاح",
    }


# ──────────────────────────────────────────────────────────────
# P8.4 — IMPORT TEMPLATE
# ──────────────────────────────────────────────────────────────

@router.get("/entities/{entity_code}/import/template",
            dependencies=[
                Depends(require_permission("dynamic", "read")),
                Depends(read_limiter.check)
            ])
async def download_import_template(
    entity_code: str,
    format: str = Query("csv", regex="^(csv|xlsx)$"),
    db: Session = Depends(get_db),
    verification: DynamicVerificationEngine = Depends(
        get_verification_engine
    ),
):
    """
    Download a CSV/Excel template for import based on entity metadata.
    """
    if not verification.entity_exists():
        raise HTTPException(
            status_code=404,
            detail=f"الكيان '{entity_code}' غير موجود"
        )

    table_error = verification.validate_table_mapping()
    if table_error:
        raise HTTPException(
            status_code=400,
            detail=table_error
        )

    # Get entity fields from metadata
    engine = MetadataEngine(db)
    schema = engine.get_full_schema(entity_code)
    fields = schema.get("fields", [])

    # Build column headers from field definitions
    headers = []
    for field in fields:
        headers.append(field["code"])

    # Add tenant_id hint for SCOPED entities
    if verification.has_tenant_id_column():
        headers.append("tenant_id")

    if format == "csv":
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(headers)
        # Add example row
        example = ["EXAMPLE-001"] + [""] * (len(headers) - 1)
        writer.writerow(example)
        output.seek(0)

        return StreamingResponse(
            iter([output.getvalue()]),
            media_type="text/csv",
            headers={
                "Content-Disposition": f"attachment; filename={entity_code}_import_template.csv"
            }
        )

    else:  # xlsx
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = entity_code

        # Write headers
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = cell.font.copy(bold=True)

        # Write example row
        ws.cell(row=2, column=1, value="EXAMPLE-001")

        # Auto-fit column widths
        for col_idx, header in enumerate(headers, 1):
            ws.column_dimensions[chr(64 + col_idx) if col_idx <= 26 else "A"].width = max(15, len(header) + 4)

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        return StreamingResponse(
            buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename={entity_code}_import_template.xlsx"
            }
        )


# ──────────────────────────────────────────────────────────────
# P8.4 — IMPORT CSV/EXCEL
# ──────────────────────────────────────────────────────────────

@router.post("/entities/{entity_code}/import",
             dependencies=[
                 Depends(require_permission("dynamic", "create")),
                 Depends(write_limiter.check)
             ])
async def import_records(
    entity_code: str,
    file: UploadFile = File(...),
    mode: str = Query("atomic", regex="^(atomic|partial)$"),
    request: Request = None,
    db: Session = Depends(get_db),
    verification: DynamicVerificationEngine = Depends(
        get_verification_engine
    ),
    current_user: Optional[dict] = Depends(optional_get_current_user)
):
    """
    Import records from CSV or Excel file.

    Modes:
        - atomic: all records must succeed, or all rollback
        - partial: successful records are saved, failed ones are reported

    Returns:
        Summary with success/error counts and per-row errors.
    """
    if not verification.entity_exists():
        raise HTTPException(
            status_code=404,
            detail=f"الكيان '{entity_code}' غير موجود"
        )

    table_error = verification.validate_table_mapping()
    if table_error:
        raise HTTPException(
            status_code=400,
            detail=table_error
        )

    # Determine file type
    filename = file.filename or ""
    is_xlsx = filename.endswith(".xlsx")
    is_csv = filename.endswith(".csv")

    if not is_xlsx and not is_csv:
        return create_error_response(
            status_code=400,
            code=ErrorCodes.VALIDATION_ERROR,
            details=[{"message": "الملف يجب أن يكون CSV أو XLSX",
                       "message_en": "File must be CSV or XLSX"}]
        )

    # Read file content
    content = await file.read()
    if len(content) == 0:
        return create_error_response(
            status_code=400,
            code=ErrorCodes.VALIDATION_ERROR,
            details=[{"message": "الملف فارغ",
                       "message_en": "File is empty"}]
        )

    # Parse rows from file
    rows: List[Dict[str, str]] = []
    headers: List[str] = []

    try:
        if is_csv:
            text_content = content.decode("utf-8-sig")
            reader = csv.DictReader(io.StringIO(text_content))
            headers = reader.fieldnames or []
            for row in reader:
                rows.append(dict(row))
        else:
            from openpyxl import load_workbook
            wb = load_workbook(io.BytesIO(content), read_only=True)
            ws = wb.active

            # Read headers from first row
            header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
            headers = [str(h).strip() if h else "" for h in header_row]

            # Read data rows
            for row in ws.iter_rows(min_row=2, values_only=True):
                row_dict = {}
                for col_idx, value in enumerate(row):
                    if col_idx < len(headers) and headers[col_idx]:
                        row_dict[headers[col_idx]] = str(value) if value is not None else ""
                rows.append(row_dict)

            wb.close()

    except Exception as e:
        return create_error_response(
            status_code=400,
            code=ErrorCodes.VALIDATION_ERROR,
            details=[{"message": f"خطأ في قراءة الملف: {str(e)}",
                       "message_en": f"Error reading file: {str(e)}"}]
        )

    if len(rows) == 0:
        return create_error_response(
            status_code=400,
            code=ErrorCodes.VALIDATION_ERROR,
            details=[{"message": "لا توجد صفوف بيانات في الملف",
                       "message_en": "No data rows in file"}]
        )

    if len(rows) > BULK_MAX_RECORDS:
        return create_error_response(
            status_code=400,
            code=ErrorCodes.VALIDATION_ERROR,
            details=[{"message": f"الحد الأقصى {BULK_MAX_RECORDS} سجل في الطلب",
                       "message_en": f"Maximum {BULK_MAX_RECORDS} records per request"}]
        )

    table_name = verification.entity_meta["table_mapping"]
    pk_col = verification.get_pk_column()

    # Determine effective tenant
    effective_tenant = None
    if verification.has_tenant_id_column():
        if not current_user:
            raise HTTPException(
                status_code=401,
                detail="Authentication required for SCOPED entity"
            )
        effective_tenant = current_user["tenant_id"]

    # Validate and prepare all rows
    collected_errors: List[Dict[str, Any]] = []
    clean_records: List[Dict[str, Any]] = []

    for idx, row in enumerate(rows):
        # Skip empty/example rows
        first_val = list(row.values())[0] if row else ""
        if first_val == "EXAMPLE-001" or all(v == "" for v in row.values()):
            continue

        # Filter to valid columns
        record_data = {}
        for key, value in row.items():
            if not key or key == pk_col or key == "tenant_id":
                continue
            if verification.check_column_exists(key):
                # Type coercion
                clean_val = value.strip() if value else None
                if clean_val == "":
                    clean_val = None
                record_data[key] = clean_val

        # Schema validation
        try:
            verification.validate_required_fields(record_data)
            verification.validate_enum_fields(record_data)
        except ValueError as e:
            collected_errors.append({
                "row": idx + 2,  # +2 for 1-indexed + header row
                "code": ErrorCodes.VALIDATION_ERROR,
                "message": str(e),
                "message_en": str(e),
            })
            continue

        # NOT NULL validation (exclude tenant_id — auto-populated by system)
        not_null_errors = verification.validate_not_null_columns(
            record_data,
            exclude_cols=["tenant_id"]
        )
        if not_null_errors:
            collected_errors.append({
                "row": idx + 2,
                "code": ErrorCodes.NOT_NULL_VIOLATION,
                "message": "; ".join(not_null_errors),
                "message_en": "; ".join(not_null_errors),
            })
            continue

        # Duplicate check
        dup_errors = verification.check_duplicate_by_unique_fields(
            record_data,
            tenant_id=effective_tenant
        )
        if dup_errors:
            collected_errors.append({
                "row": idx + 2,
                "code": ErrorCodes.DUPLICATE_RECORD,
                "message": "; ".join(dup_errors),
                "message_en": "; ".join(dup_errors),
            })
            continue

        # Build clean data
        clean_data = {pk_col: verification.generate_pk_value()}
        if verification.has_tenant_id_column():
            clean_data["tenant_id"] = effective_tenant

        for key, value in record_data.items():
            clean_data[key] = value

        clean_records.append(clean_data)

    # Handle based on mode
    if mode == "atomic" and collected_errors:
        return create_error_response(
            status_code=400,
            code=ErrorCodes.VALIDATION_ERROR,
            details=[{"message": f"تم العثور على {len(collected_errors)} أخطاء",
                       "message_en": f"Found {len(collected_errors)} errors",
                       "errors": collected_errors}]
        )

    # Execute imports
    created_ids = []
    partial_errors = list(collected_errors) if mode == "partial" else []

    try:
        for clean_data in clean_records:
            cols = ", ".join(clean_data.keys())
            placeholders = ", ".join([f":{k}" for k in clean_data.keys()])

            query = text(
                f"""
                INSERT INTO {table_name}
                ({cols})
                VALUES ({placeholders})
                RETURNING {pk_col}
                """
            )

            result = db.execute(query, clean_data)
            new_id = result.scalar()
            created_ids.append(new_id)

            # Per-record audit
            log_dynamic_audit(
                db=db,
                tenant_id=effective_tenant or "unknown",
                user_id=current_user.get("id") if current_user else None,
                user_email=current_user.get("email") if current_user else None,
                action="import",
                entity_code=entity_code,
                record_id=new_id,
                new_values=clean_data,
                ip_address=request.client.host if request and request.client else None,
                user_agent=request.headers.get("user-agent") if request else None,
                status="success",
            )

        EventBus(db).emit("record.imported", entity_code, tenant_id=effective_tenant, user_id=current_user.get("id") if current_user else None, payload={"imported_count": len(created_ids), "error_count": len(partial_errors)})

        db.commit()

        return {
            "status": "success",
            "imported": len(created_ids),
            "errors": len(partial_errors),
            "error_details": partial_errors if mode == "partial" else [],
            "tenant_capability": verification.tenant_capability,
            "effective_tenant": effective_tenant,
        }

    except Exception as e:
        db.rollback()

        log_dynamic_audit(
            db=db,
            tenant_id=effective_tenant or "unknown",
            user_id=current_user.get("id") if current_user else None,
            user_email=current_user.get("email") if current_user else None,
            action="import",
            entity_code=entity_code,
            new_values={"attempted_count": len(clean_records)},
            ip_address=request.client.host if request and request.client else None,
            user_agent=request.headers.get("user-agent") if request else None,
            status="failure",
            error_message=str(e),
        )

        return secure_db_error(e)


# ──────────────────────────────────────────────────────────────
# P8.5 — EXPORT CSV/EXCEL
# ──────────────────────────────────────────────────────────────

EXPORT_MAX_ROWS = 10000


@router.get("/entities/{entity_code}/export",
            dependencies=[
                Depends(require_permission("dynamic", "read")),
                Depends(read_limiter.check)
            ])
async def export_records(
    entity_code: str,
    format: str = Query("csv", regex="^(csv|xlsx)$"),
    filters: Optional[str] = None,
    sort: Optional[str] = None,
    limit: int = 100,
    db: Session = Depends(get_db),
    verification: DynamicVerificationEngine = Depends(
        get_verification_engine
    ),
    current_user: Optional[dict] = Depends(optional_get_current_user)
):
    """
    Export records to CSV or Excel.

    Respects:
        - Tenant scope (SCOPED = only auth tenant's data)
        - P7 Query Engine (filters, sorting)
        - Max 10000 rows
    """
    if not verification.entity_exists():
        raise HTTPException(
            status_code=404,
            detail=f"الكيان '{entity_code}' غير موجود"
        )

    table_error = verification.validate_table_mapping()
    if table_error:
        raise HTTPException(
            status_code=400,
            detail=table_error
        )

    # Parse and validate query parameters
    try:
        parser = QueryParser(verification.real_columns)
        query_filter = parser.parse_query(
            filters_str=filters,
            sort_str=sort,
            limit=min(limit, EXPORT_MAX_ROWS),
            offset=0
        )
    except QueryParseError as e:
        return create_error_response(
            status_code=400,
            code=ErrorCodes.VALIDATION_ERROR,
            details=[{"message": str(e), "message_en": str(e)}]
        )

    table_name = verification.entity_meta["table_mapping"]
    where_clauses = []
    params = {}

    # Tenant scope (NEVER from user filters)
    if verification.has_tenant_id_column():
        if not current_user:
            raise HTTPException(
                status_code=401,
                detail="Authentication required for SCOPED entity"
            )
        auth_tenant_id = current_user["tenant_id"]
        where_clauses.append("tenant_id = :tenant_id")
        params["tenant_id"] = auth_tenant_id

    # P10: Exclude soft-deleted records from export
    has_deleted_export = verification.real_columns.get("deleted_at") is not None
    if has_deleted_export:
        where_clauses.append("deleted_at IS NULL")

    # Add user filters
    user_where, user_params = parser.build_where_clause(query_filter.filters)
    if user_where:
        where_clauses.append(user_where)
        params.update(user_params)

    where_sql = " AND ".join(where_clauses) if where_clauses else "1=1"

    # Build ORDER BY
    order_sql = parser.build_order_clause(query_filter.sorts)
    if not order_sql:
        order_sql = "ORDER BY created_at DESC"

    # Get total count
    count_query = text(f"SELECT COUNT(*) FROM {table_name} WHERE {where_sql}")
    total = db.execute(count_query, params).scalar()

    # Get records
    limit = query_filter.limit
    query = text(
        f"SELECT * FROM {table_name} "
        f"WHERE {where_sql} "
        f"{order_sql} "
        f"LIMIT :limit OFFSET :offset"
    )
    params["limit"] = limit
    params["offset"] = 0

    result = db.execute(query, params)
    columns = [col for col in result.keys() if col != "tenant_id"]
    rows_data = [dict(row._mapping) for row in result]

    # Build output
    if format == "csv":
        output = io.StringIO()
        writer = csv.DictWriter(output, fieldnames=columns, extrasaction='ignore')
        writer.writeheader()
        for row in rows_data:
            clean_row = {k: v for k, v in row.items() if k != "tenant_id"}
            writer.writerow(clean_row)
        output.seek(0)

        return StreamingResponse(
            iter([output.getvalue()]),
            media_type="text/csv",
            headers={
                "Content-Disposition": f"attachment; filename={entity_code}_export.csv",
                "X-Export-Total": str(total),
                "X-Export-Count": str(len(rows_data)),
            }
        )

    else:  # xlsx
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = entity_code

        # Write headers
        for col_idx, header in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = cell.font.copy(bold=True)

        # Write data
        for row_idx, row in enumerate(rows_data, 2):
            for col_idx, col_name in enumerate(columns, 1):
                ws.cell(row=row_idx, column=col_idx, value=row.get(col_name))

        # Auto-fit
        for col_idx, header in enumerate(columns, 1):
            letter = chr(64 + col_idx) if col_idx <= 26 else "A"
            ws.column_dimensions[letter].width = max(15, len(header) + 4)

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        return StreamingResponse(
            buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename={entity_code}_export.xlsx",
                "X-Export-Total": str(total),
                "X-Export-Count": str(len(rows_data)),
            }
        )
